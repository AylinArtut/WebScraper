from datetime import timedelta, datetime, date as dt_date, date
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
import json
import re
import os

def load_config(file_path):
    with open(file_path) as f:
        config = json.load(f)
        return (
            config['credentials'],
            list(config['forum_urls'].values()),
            list(config['exercise_urls'].values()),
            config['calendar_weeks'],
            config['excluded_usernames'],
            config['multipleChoice_tests']
        )

config_file_path = 'config.json'
(credentials, forum_urls, exercise_urls, calendar_weeks, excluded_usernames,
 multipleChoice_tests) = load_config(config_file_path)

username = credentials['username']
password = credentials['password']
authentication = credentials['authentication']
login_url = credentials['login_url']
request_url = credentials['request_url']
excel_path = credentials['excel_path']
multipleChoice_json = credentials['multipleChoice_json']
uebungDownload_json = credentials['uebungDownload_json']
config_json = credentials['config_json']
url_forumUebersicht = credentials['url_forumUebersicht']

def normalize_date_format(date_str):
    month_mapping = {
        'Jan': '01', 'Feb': '02', 'Mär': '03', 'Apr': '04', 'Mai': '05', 'Jun': '06',
        'Jul': '07', 'Aug': '08', 'Sep': '09', 'Okt': '10', 'Nov': '11', 'Dez': '12'
    }
    parts = date_str.split()
    if parts[1] in month_mapping:
        parts[1] = month_mapping[parts[1]]
    processed_date_str = ' '.join(parts).split(',')[0].strip()
    return processed_date_str

def parse_post_info(response_text):
    soup = BeautifulSoup(response_text, 'html.parser')
    post_headers = soup.find_all('div', {'class': 'ilFrmPostHeader'})
    titles_dates_usernames = []
    for header in post_headers:
        post_info = header.find('span', {'class': 'small'}).text.strip()
        post_info_parts = post_info.split('|')
        date_str = post_info_parts[-1].strip()
        username = post_info_parts[0].strip()
        date_time_obj = datetime.now() if "Heute" in date_str else (
            (datetime.now() - timedelta(days=1))
            if "Gestern" in date_str
            else datetime.strptime(normalize_date_format(date_str), '%d. %m %Y')
        )
        title = header.find('div', {'class': 'ilFrmPostTitle'}).text.strip()
        titles_dates_usernames.append((title, date_time_obj, username))
    return titles_dates_usernames

def addData_to_jsonFile(data_to_add, file_path):
    if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                content = json.load(file)
        except json.JSONDecodeError:
            content = []
    else:
        content = []
    entry = {key.isoformat(): value for key, value in data_to_add.items()}
    if entry not in content:
        content.append(entry)
        with open(file_path, "w", encoding="utf-8") as file:
            json.dump(content, file, ensure_ascii=False, indent=4)

def fetch_Readers(url, session, exercise_number):
    download_und_datum = {}
    main_page_response = session.get(url)
    main_page_soup = BeautifulSoup(main_page_response.text, 'html.parser')
    pdf_links = main_page_soup.find_all('a', {'class': 'il_ContainerItemTitle'})
    true_link = next((link for link in pdf_links if "Übung" in link.text), None)
    if true_link:
        pdf_url = true_link['href']
        match = re.search(r'\d+', pdf_url)
        if match:
            first_number = match.group()
            new_url = f"{credentials['url_to_work_with']}/goto.php?target=fold_{first_number}&client_id=db_040811"
            new_page_response = session.get(new_url)
            new_page_soup = BeautifulSoup(new_page_response.text, 'html.parser')
            page_text = new_page_soup.get_text()
            match = re.search(r'Gelesen von LEA-Benutzern \(Anzahl\)\s*(\d+)', page_text)
            if match:
                number_of_readers = match.group(1)
                today = date.today()
                download_und_datum[today] = {'number_of_readers': number_of_readers,
                                             'exercise_number': exercise_number - 1,
                                             'calendar_week': determine_week(today, calendar_weeks)}
                addData_to_jsonFile(download_und_datum, uebungDownload_json)
            else:
                print(f"Übung {exercise_number - 1}, gelesen von LEA-Benutzern (Anzahl): 0")
        else:
            print(f"Keine Zahlenfolge gefunden für Übung {exercise_number - 1}.")
    else:
        print(f"PDF-Datei nicht gefunden für Übung {exercise_number - 1}.")
    return [(data['number_of_readers'], data['exercise_number'], data.get('calendar_week'))
            for data in download_und_datum.values()]

def determine_week(datum, calendar_weeks):
    for week in calendar_weeks:
        start_date = datetime.strptime(week[1], "%d.%m.%y").date()
        end_date = datetime.strptime(week[2], "%d.%m.%y").date()
        if start_date <= datum <= end_date:
            return week[0]
    return None

def scrape_postCounts(session, forum_urls):
    week_title_counts = {week: 0 for week, _, _ in calendar_weeks}
    for url in forum_urls:
        response = session.get(url)
        titles_and_dates = parse_post_info(response.text)
        for title, date, username in titles_and_dates:
            for week, date1, date2 in calendar_weeks:
                date1_obj = datetime.strptime(date1, '%d.%m.%y')
                date2_obj = datetime.strptime(date2, '%d.%m.%y')
                if date1_obj <= date <= date2_obj or date2_obj <= date <= date1_obj:
                    excluded = any(username == excluded_user[1] or
                                   username == excluded_user[0] for excluded_user in excluded_usernames)
                    if not excluded:
                        week_title_counts[week] += 1
    return week_title_counts

def main():
    with open(config_json, 'r') as file:
        data = json.load(file)
    with requests.Session() as session:
        session.post(login_url, data={'username': username, 'password': password,
                                      'cmd[doStandardAuthentication]': authentication}, allow_redirects=True)
        response = session.get(url_forumUebersicht)
        soup_forumUebersicht = BeautifulSoup(response.text, 'html.parser')
        htmlTag_forumUebersicht = soup_forumUebersicht.find_all('td', class_='std small')
        new_links = {}
        for td_tag in htmlTag_forumUebersicht:
            link = td_tag.find('a')
            if link:
                link_href = f"{credentials['url_to_work_with']}/{link['href']}"
                if link_href not in data['forum_urls'].values():
                    key = str(len(data['forum_urls']) + len(new_links) + 2)
                    new_links[key] = link_href
        data['forum_urls'].update(new_links)
    with open(config_json, 'w') as file:
        json.dump(data, file, indent=4)
        for index, exercise_url in enumerate(exercise_urls, 1):
            fetch_Readers(exercise_url, session, index)
        week_counts = scrape_postCounts(session, forum_urls)
    calendar_weeks_only = [week[0] for week in calendar_weeks]
    book = load_workbook(excel_path)
    sheet = book.active
    sheet['A2'].value = "Anzahl Forum-Fragen"
    column_letters = 'BCDEFGHIJKLMNOPQRSTUVWXYZ'
    for i, (week, count) in enumerate(week_counts.items(), start=2):
        sheet.cell(row=2, column=i, value=count)
    for i in range(1, 11):
        sheet[f'A{i + 2}'].value = f"Download Übungsblatt {i - 1}"
    for i in range(13, 26):
        sheet[f'A{i}'].value = f"Selbsttest {i - 12}"
    for i, week in enumerate(calendar_weeks_only):
        sheet[column_letters[i] + '1'].value = week
    columns = {'KW14': 'B', 'KW15': 'C', 'KW16': 'D', 'KW17': 'E', 'KW18': 'F', 'KW19': 'G', 'KW20': 'H', 'KW21':
               'I', 'KW22': 'J', 'KW23': 'K', 'KW24': 'L', 'KW25': 'M', 'KW26': 'N', 'KW27': 'O', 'KW28': 'P'}
    with open(uebungDownload_json, 'r') as file:
        histories = json.load(file)
        for entry in histories:
            date = list(entry.keys())[0]
            data = entry[date]
            number_of_readers = data["number_of_readers"]
            exercise_number = data["exercise_number"]
            calendar_week = data["calendar_week"]
            for j in range(3, 13):
                if sheet[f'A{j}'].value == f'Download Übungsblatt {exercise_number}':
                    sheet[f'{columns[f"{calendar_week}"]}{j}'].value = number_of_readers
    for test_name, test_id in multipleChoice_tests.items():
        post_data = {"get": {"app": f"{test_id}"}, "store": "mc-results"}
        reply = requests.post(f'{request_url}', json=post_data)
        obj = reply.json()
        user_list = list()
        questions = dict()
        for user_submission in obj:
            if user_submission['user'] not in user_list:
                user_list.append(user_submission['user'])
            for question in user_submission['questions']:
                if question['key'] not in questions:
                    questions[question['key']] = dict()
                    questions[question['key']]['text'] = question['text']
                    questions[question['key']]['count'] = 0
                questions[question['key']]['count'] += 1
        sumCounts = sum(value['count'] for value in questions.values()) / 5
        multipleChoice_test_und_datum = {}
        today2 = dt_date.today()
        multipleChoice_test_und_datum[today2] = {'selbsttest': test_name, 'count': int(sumCounts),
                                                 'calendar_week': determine_week(today2, calendar_weeks)}
        addData_to_jsonFile(multipleChoice_test_und_datum, multipleChoice_json)
        with open(multipleChoice_json, 'r') as file:
            histories = json.load(file)
            for entry in histories:
                date = list(entry.keys())[0]
                data = entry[date]
                selbsttest = data["selbsttest"]
                count = data["count"]
                calendar_week = data["calendar_week"]
                for j in range(13, 26):
                    if sheet[f'A{j}'].value == f'{selbsttest}':
                        sheet[f'{columns[f"{calendar_week}"]}{j}'].value = count
    book.save(excel_path)

if __name__ == "__main__":
    main()
